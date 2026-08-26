"""Parsing, normalization, and find-or-create helpers for the Excel bulk-import feature
(Interviews page "Import" button). Kept separate from the router so the parsing logic can be
exercised without spinning up a request.

Column layouts and the two documented column-shift blocks were verified directly against
`Local Interviews.xlsx` (AIML rows 302-308, DevOps rows 1-6) — see the import plan for details.
"""

from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from typing import Iterator, Optional

from dateutil import parser as dateutil_parser
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from app.models.business_developer import BusinessDeveloper
from app.models.candidate import Candidate
from app.models.company import Company
from app.models.resume_profile import ResumeProfile

# ─── Text / key normalization ──────────────────────────────────────────────


def normalize_text(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def normalize_key(value) -> Optional[str]:
    s = normalize_text(value)
    return s.lower() if s else None


# ─── Round-label normalization ─────────────────────────────────────────────

_ROUND_SUFFIX = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th", 6: "6th"}
_ROUND_DIGIT_RE = re.compile(r"^\s*(\d+)")


def normalize_round(value) -> str:
    """Best-effort round-label cleanup. Never returns empty — unparseable text passes through
    as literal rather than being treated as a skip reason (round label alone is low-stakes)."""
    if value is None:
        return "Unspecified"
    if isinstance(value, (int, float)):
        n = int(value)
        return _ROUND_SUFFIX.get(n, f"{n}th")
    s = normalize_text(value)
    if not s:
        return "Unspecified"
    if s.lower().startswith("final"):
        return "Final"
    m = _ROUND_DIGIT_RE.match(s)
    if m:
        n = int(m.group(1))
        return _ROUND_SUFFIX.get(n, f"{n}th")
    return s


# ─── Date parsing ───────────────────────────────────────────────────────────

_EXCEL_EPOCH = date(1899, 12, 30)
_DATE_FORMATS = [
    "%d %b %Y", "%d %B %Y", "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y",
]
_SANE_MIN, _SANE_MAX = date(2015, 1, 1), date(2035, 12, 31)


def parse_date_cell(value) -> tuple[Optional[date], Optional[str]]:
    """Returns (parsed_date, error_reason). Never raises — callers skip-and-report on error."""
    if value is None:
        return None, "missing date"
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    if isinstance(value, time):
        return None, "date cell contains a time-only value with no date component"
    if isinstance(value, (int, float)):
        d = _EXCEL_EPOCH + timedelta(days=float(value))
        if not (_SANE_MIN <= d <= _SANE_MAX):
            return None, f"excel serial date out of sane range: {value!r} -> {d}"
        return d, None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None, "missing date"
        # Fixes a real data-entry typo: '14 APR 2O26' (letter O instead of zero).
        s_fixed = re.sub(r"(\d)O(\d)", r"\g<1>0\g<2>", s)
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(s_fixed, fmt).date(), None
            except ValueError:
                continue
        try:
            d = dateutil_parser.parse(s_fixed, fuzzy=True, dayfirst=False).date()
        except (ValueError, OverflowError, TypeError):
            return None, f"unparseable date string: {value!r}"
        if not (_SANE_MIN <= d <= _SANE_MAX):
            return None, f"parsed date out of sane range: {value!r} -> {d}"
        return d, None
    return None, f"unrecognized date cell type: {type(value).__name__}"


# ─── Time parsing (best-effort, never blocks a row) ────────────────────────

_TIME_RE = re.compile(r"(\d{1,2})(?::(\d{2}))?\s*([ap])\.?\s*m?\.?", re.I)


def parse_time_cell(value) -> Optional[time]:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        first_leg = re.split(r"[-–—]", value.strip())[0]
        m = _TIME_RE.search(first_leg)
        if not m:
            return None
        h, mnt, ap = int(m.group(1)), int(m.group(2) or 0), m.group(3).lower()
        if h == 12:
            h = 0
        if ap == "p":
            h += 12
        if 0 <= h <= 23 and 0 <= mnt <= 59:
            return time(hour=h, minute=mnt)
    return None


# ─── Find-or-create helpers (case-insensitive name match, in-memory cache) ─
# The importer's own cache is pre-loaded once (see load_*_cache below) and shared across all
# three sheets, so the same name mentioned in two sheets resolves to one DB row. Department is
# stamped only when a row is newly created here — an existing match keeps whatever department
# it already has, so re-running the import never silently overwrites a manual edit made later
# in the live app.


def load_company_cache(session: Session) -> dict[str, Company]:
    return {c.name.lower(): c for c in session.exec(select(Company)).all()}


def load_candidate_cache(session: Session) -> dict[str, Candidate]:
    return {c.name.lower(): c for c in session.exec(select(Candidate)).all()}


def load_resume_profile_cache(session: Session) -> dict[str, ResumeProfile]:
    return {p.name.lower(): p for p in session.exec(select(ResumeProfile)).all()}


def load_business_developer_cache(session: Session) -> dict[str, BusinessDeveloper]:
    return {b.name.lower(): b for b in session.exec(select(BusinessDeveloper)).all()}


def get_or_create_company(session: Session, cache: dict[str, Company], name_raw) -> Optional[Company]:
    name = normalize_text(name_raw)
    if not name:
        return None
    key = name.lower()
    if key in cache:
        return cache[key]
    company = Company(name=name)
    session.add(company)
    session.flush()
    cache[key] = company
    return company


def get_or_create_candidate(
    session: Session, cache: dict[str, Candidate], name_raw, department_id: Optional[uuid.UUID]
) -> Optional[Candidate]:
    name = normalize_text(name_raw)
    if not name:
        return None
    key = name.lower()
    if key in cache:
        return cache[key]
    candidate = Candidate(name=name, department_id=department_id)
    session.add(candidate)
    session.flush()
    cache[key] = candidate
    return candidate


def get_or_create_resume_profile(
    session: Session, cache: dict[str, ResumeProfile], name_raw, department_id: Optional[uuid.UUID]
) -> Optional[ResumeProfile]:
    name = normalize_text(name_raw)
    if not name:
        return None
    key = name.lower()
    if key in cache:
        return cache[key]
    profile = ResumeProfile(name=name, department_id=department_id)
    session.add(profile)
    session.flush()
    cache[key] = profile
    return profile


def get_or_create_business_developer(
    session: Session, cache: dict[str, BusinessDeveloper], name_raw
) -> Optional[BusinessDeveloper]:
    name = normalize_text(name_raw)
    if not name:
        return None
    key = name.lower()
    if key in cache:
        return cache[key]
    bd = BusinessDeveloper(name=name)
    session.add(bd)
    session.flush()
    cache[key] = bd
    return bd


# ─── Parsed row shape ───────────────────────────────────────────────────────


@dataclass
class ParsedRow:
    sheet: str
    excel_row: int
    status_raw: Optional[str]
    bd_name: Optional[str]
    candidate_name: Optional[str]
    company_name: Optional[str]
    position: Optional[str]
    round_label: str
    interviewer: Optional[str]
    interview_date: Optional[date]
    date_error: Optional[str]
    time_est: Optional[time]
    dev_feedback: Optional[str]
    client_feedback: Optional[str]
    barriers: Optional[str]


def dedup_key(row: ParsedRow) -> tuple:
    """In-file duplicate key. Includes position + date (not just candidate/company/round) since
    the same candidate legitimately interviews multiple times at the same staffing company for
    different end-client roles/dates."""
    return (
        normalize_key(row.candidate_name),
        normalize_key(row.company_name),
        row.round_label.strip().lower(),
        normalize_key(row.position),
        row.interview_date,
    )


# ─── Sheet stop condition ───────────────────────────────────────────────────


def last_real_row(ws: Worksheet, min_col: int = 2) -> int:
    """Nominal ws.max_row on these sheets is mostly empty padding — find the true last row
    with any data beyond the row-number column."""
    last = 0
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column):
        row_idx = row[0].row
        if any(normalize_text(c.value) is not None for c in row[min_col - 1:]):
            last = row_idx
    return last


# ─── Per-sheet readers ──────────────────────────────────────────────────────

_ROUND_LIKE_RE = re.compile(r"^\s*(\d+(st|nd|rd|th)?|final)\b", re.I)


def read_aiml_sheet(ws: Worksheet) -> Iterator[ParsedRow]:
    last = last_real_row(ws)
    for row in ws.iter_rows(min_row=2, max_row=last, values_only=False):
        excel_row = row[0].row
        vals = [c.value for c in row]
        # 1-indexed per the documented mapping; vals is 0-indexed.
        status_raw = vals[1] if len(vals) > 1 else None
        bd_name = vals[2] if len(vals) > 2 else None
        candidate_name = vals[3] if len(vals) > 3 else None
        company = vals[5] if len(vals) > 5 else None
        position = vals[6] if len(vals) > 6 else None
        pay_scale = vals[7] if len(vals) > 7 else None
        round_val = vals[8] if len(vals) > 8 else None
        interviewer = vals[9] if len(vals) > 9 else None
        date_val = vals[11] if len(vals) > 11 else None
        time_val = vals[12] if len(vals) > 12 else None
        dev_feedback = vals[13] if len(vals) > 13 else None
        client_feedback = vals[14] if len(vals) > 14 else None
        barriers = vals[15] if len(vals) > 15 else None

        # Verified data-entry error block (AIML rows 302-308): columns shifted one slot left
        # starting at Company. Detected heuristically — Position empty AND the cell that would
        # normally hold Pay Scale instead looks like a round label.
        if normalize_text(position) is None and isinstance(pay_scale, str) and _ROUND_LIKE_RE.match(pay_scale.strip()):
            company = vals[4] if len(vals) > 4 else None
            position = vals[5] if len(vals) > 5 else None
            round_val = pay_scale
            interviewer = vals[8] if len(vals) > 8 else None
            date_val = vals[10] if len(vals) > 10 else None
            time_val = vals[11] if len(vals) > 11 else None

        d, date_err = parse_date_cell(date_val)
        yield ParsedRow(
            sheet="AIML",
            excel_row=excel_row,
            status_raw=normalize_text(status_raw),
            bd_name=normalize_text(bd_name),
            candidate_name=normalize_text(candidate_name),
            company_name=normalize_text(company),
            position=normalize_text(position),
            round_label=normalize_round(round_val),
            interviewer=normalize_text(interviewer),
            interview_date=d,
            date_error=date_err,
            time_est=parse_time_cell(time_val),
            dev_feedback=normalize_text(dev_feedback),
            client_feedback=normalize_text(client_feedback),
            barriers=normalize_text(barriers),
        )


# Verified fixed range: DevOps rows 1-6 are shifted one column left from Status onward.
_DEVOPS_SHIFTED_ROWS = {1, 2, 3, 4, 5, 6}


def read_devops_sheet(ws: Worksheet) -> Iterator[ParsedRow]:
    last = last_real_row(ws)
    for row in ws.iter_rows(min_row=1, max_row=last, values_only=False):
        excel_row = row[0].row
        vals = [c.value for c in row]

        if excel_row in _DEVOPS_SHIFTED_ROWS:
            status_raw = None  # col2 holds the BD's name here, never a real status
            bd_name = vals[1] if len(vals) > 1 else None
            candidate_name = vals[2] if len(vals) > 2 else None
            company = vals[3] if len(vals) > 3 else None
            position = vals[4] if len(vals) > 4 else None
            round_val = vals[6] if len(vals) > 6 else None
            interviewer = vals[7] if len(vals) > 7 else None
            date_val = vals[9] if len(vals) > 9 else None
            time_val = vals[10] if len(vals) > 10 else None
            dev_feedback = vals[11] if len(vals) > 11 else None
            client_feedback = vals[12] if len(vals) > 12 else None
            barriers = vals[13] if len(vals) > 13 else None
            if excel_row == 1:
                # Row 1 has literal leftover header text ('Time', "Developer's feedback",
                # 'Client feedback') in these slots post-shift — not real data.
                dev_feedback = None
                client_feedback = None
                barriers = None
        else:
            status_raw = vals[1] if len(vals) > 1 else None
            bd_name = vals[2] if len(vals) > 2 else None
            candidate_name = vals[3] if len(vals) > 3 else None
            company = vals[4] if len(vals) > 4 else None
            position = vals[5] if len(vals) > 5 else None
            round_val = vals[7] if len(vals) > 7 else None
            interviewer = vals[8] if len(vals) > 8 else None
            date_val = vals[10] if len(vals) > 10 else None
            time_val = vals[11] if len(vals) > 11 else None
            dev_feedback = vals[12] if len(vals) > 12 else None
            client_feedback = vals[13] if len(vals) > 13 else None
            barriers = vals[14] if len(vals) > 14 else None

        d, date_err = parse_date_cell(date_val)
        yield ParsedRow(
            sheet="DevOps",
            excel_row=excel_row,
            status_raw=normalize_text(status_raw),
            bd_name=normalize_text(bd_name),
            candidate_name=normalize_text(candidate_name),
            company_name=normalize_text(company),
            position=normalize_text(position),
            round_label=normalize_round(round_val),
            interviewer=normalize_text(interviewer),
            interview_date=d,
            date_error=date_err,
            time_est=parse_time_cell(time_val),
            dev_feedback=normalize_text(dev_feedback),
            client_feedback=normalize_text(client_feedback),
            barriers=normalize_text(barriers),
        )


def read_data_sheet(ws: Worksheet) -> Iterator[ParsedRow]:
    last = last_real_row(ws)
    for row in ws.iter_rows(min_row=2, max_row=last, values_only=False):
        excel_row = row[0].row
        vals = [c.value for c in row]
        status_raw = vals[1] if len(vals) > 1 else None
        bd_name = vals[2] if len(vals) > 2 else None
        candidate_name = vals[3] if len(vals) > 3 else None
        company = vals[4] if len(vals) > 4 else None
        position = vals[5] if len(vals) > 5 else None
        round_val = vals[7] if len(vals) > 7 else None
        interviewer = vals[8] if len(vals) > 8 else None  # labeled "Recruiter" in this sheet
        # col 9 ("Interview Scheduling Date") is informational only — not stored on Interview.
        date_val = vals[10] if len(vals) > 10 else None
        time_val = vals[11] if len(vals) > 11 else None
        dev_feedback = vals[12] if len(vals) > 12 else None

        d, date_err = parse_date_cell(date_val)
        yield ParsedRow(
            sheet="Data",
            excel_row=excel_row,
            status_raw=normalize_text(status_raw),
            bd_name=normalize_text(bd_name),
            candidate_name=normalize_text(candidate_name),
            company_name=normalize_text(company),
            position=normalize_text(position),
            round_label=normalize_round(round_val),
            interviewer=normalize_text(interviewer),
            interview_date=d,
            date_error=date_err,
            time_est=parse_time_cell(time_val),
            dev_feedback=normalize_text(dev_feedback),
            client_feedback=None,
            barriers=None,
        )


SHEET_READERS = {
    "AIML": read_aiml_sheet,
    "DevOps": read_devops_sheet,
    "Data": read_data_sheet,
}
