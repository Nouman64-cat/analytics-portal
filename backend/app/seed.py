"""
Seed script: Imports interview data from the Excel spreadsheet into PostgreSQL.

Usage:
    cd backend
    source .venv/bin/activate
    python -m app.seed
"""

import sys
import os
from datetime import datetime

import openpyxl
from sqlmodel import Session, select

# Ensure app module is importable
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import engine, create_db_and_tables
from app.models.candidate import Candidate
from app.models.resume_profile import ResumeProfile
from app.models.company import Company
from app.models.interview import Interview


EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "Interviews Info - Nouman Ejaz - AI_ML Lead.xlsx",
)


def seed_database():
    print(f"📂 Reading Excel file: {EXCEL_PATH}")

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[wb.sheetnames[0]]

    # Create tables
    print("🗄️  Creating database tables...")
    create_db_and_tables()

    with Session(engine) as session:
        # Caches to avoid duplicates
        candidates_cache: dict[str, Candidate] = {}
        profiles_cache: dict[str, ResumeProfile] = {}
        companies_cache: dict[str, Company] = {}

        # Pre-load existing records
        for c in session.exec(select(Candidate)).all():
            candidates_cache[c.name.lower()] = c
        for p in session.exec(select(ResumeProfile)).all():
            profiles_cache[p.name.lower()] = p
        for co in session.exec(select(Company)).all():
            companies_cache[co.name.lower()] = co

        interview_count = 0
        skipped = 0

        # Row 1 is header, data starts at row 2
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            company_name = row[0]  # Col A: Agency / End Client
            staffing_firm = row[1]  # Col B: Staffing Firm
            profile_name = row[2]  # Col C: Resume Profile
            candidate_name = row[3]  # Col D: Interviewee
            role = row[4]  # Col E: Role
            salary_range = row[5]  # Col F: Salary Range
            round_val = row[6]  # Col G: Round
            interview_date = row[7]  # Col H: Date
            time_est = row[8]  # Col I: Time (EST)
            time_pkt = row[9]  # Col J: Time (PKT)
            status_val = row[10]  # Col K: Status
            feedback = row[11]  # Col L: Feedback

            # Skip empty rows
            if not company_name and not role:
                continue

            # Clean up string values
            company_name = str(company_name).strip() if company_name else None
            staffing_firm = str(staffing_firm).strip() if staffing_firm else None
            profile_name = str(profile_name).strip() if profile_name else None
            candidate_name = str(candidate_name).strip() if candidate_name else None
            role = str(role).strip() if role else "Unknown Role"
            salary_range = str(salary_range).strip() if salary_range else None
            round_val = str(round_val).strip() if round_val else "Unknown"
            status_val = str(status_val).strip() if status_val else None
            feedback = str(feedback).strip() if feedback else None

            if not company_name:
                print(f"  ⚠️  Row {row_idx}: Skipping — no company name")
                skipped += 1
                continue

            # Get or create Company
            co_key = company_name.lower()
            if co_key not in companies_cache:
                company = Company(name=company_name, staffing_firm=staffing_firm)
                session.add(company)
                session.flush()
                companies_cache[co_key] = company
                print(f"  🏢 Created company: {company_name}")
            else:
                company = companies_cache[co_key]
                # Update staffing firm if it was previously null
                if staffing_firm and not company.staffing_firm:
                    company.staffing_firm = staffing_firm
                    session.add(company)

            # Get or create Candidate
            if candidate_name:
                cand_key = candidate_name.lower()
                if cand_key not in candidates_cache:
                    candidate = Candidate(name=candidate_name)
                    session.add(candidate)
                    session.flush()
                    candidates_cache[cand_key] = candidate
                    print(f"  👤 Created candidate: {candidate_name}")
                else:
                    candidate = candidates_cache[cand_key]
            else:
                print(f"  ⚠️  Row {row_idx}: No candidate name, skipping")
                skipped += 1
                continue

            # Get or create Resume Profile
            if profile_name:
                prof_key = profile_name.lower()
                if prof_key not in profiles_cache:
                    profile = ResumeProfile(name=profile_name)
                    session.add(profile)
                    session.flush()
                    profiles_cache[prof_key] = profile
                    print(f"  📄 Created resume profile: {profile_name}")
                else:
                    profile = profiles_cache[prof_key]
            else:
                print(f"  ⚠️  Row {row_idx}: No profile name, skipping")
                skipped += 1
                continue

            # Parse date
            parsed_date = None
            if interview_date:
                if isinstance(interview_date, datetime):
                    parsed_date = interview_date.date()
                elif isinstance(interview_date, str):
                    try:
                        parsed_date = datetime.strptime(interview_date, "%Y-%m-%d").date()
                    except ValueError:
                        pass

            # Parse times
            parsed_time_est = None
            parsed_time_pkt = None
            if time_est:
                from datetime import time as time_type
                if isinstance(time_est, time_type):
                    parsed_time_est = time_est
            if time_pkt:
                from datetime import time as time_type
                if isinstance(time_pkt, time_type):
                    parsed_time_pkt = time_pkt

            # Create Interview
            interview = Interview(
                company_id=company.id,
                candidate_id=candidate.id,
                resume_profile_id=profile.id,
                role=role,
                salary_range=salary_range,
                round=round_val,
                interview_date=parsed_date,
                time_est=parsed_time_est,
                time_pkt=parsed_time_pkt,
                status=status_val,
                feedback=feedback,
            )
            session.add(interview)
            interview_count += 1

        session.commit()

        print(f"\n✅ Seeding complete!")
        print(f"   📊 Interviews created: {interview_count}")
        print(f"   🏢 Companies: {len(companies_cache)}")
        print(f"   👤 Candidates: {len(candidates_cache)}")
        print(f"   📄 Resume Profiles: {len(profiles_cache)}")
        print(f"   ⚠️  Rows skipped: {skipped}")


if __name__ == "__main__":
    seed_database()
